# import xlrd
#
# book = xlrd.open_workbook("myfile.xls")
# print("The number of worksheets is {0}".format(book.nsheets))
# print("Worksheet name(s): {0}".format(book.sheet_names()))
# sh = book.sheet_by_index(0)
# print("{0} {1} {2}".format(sh.name, sh.nrows, sh.ncols))
# print("Cell D30 is {0}".format(sh.cell_value(rowx=5, colx=0)))
# # print(sh.nrows)
# print(type(sh.row(2)[0]), sh.row(2)[0])
# print(sh.row(1)[0], type(sh.row(1)[0]), sh.row(1)[0])
# # print(sh.row(2).index(1))
# # for rx in range(sh.nrows):
# #     print(sh.row(rx))
#
# citys = []
#

# for rx in range(sh.nrows):
#     if rx != 0:
#         citys.append(sh.row(rx)[0].value)
#
# print(citys)

from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()
ws.append(['x','y','z'])

# Save the file
wb.save("sample.xlsx")

